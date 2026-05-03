import streamlit as st
import pandas as pd
import numpy as np
import io
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# --- BRANDED PAGE CONFIG ---
st.set_page_config(page_title="Sai Goud Report Processor", layout="wide")

def process_pollution_report(file_bytes):
    # 1. LOAD WORKBOOK FOR FORMATTING PRESERVATION
    file_bytes.seek(0)
    wb = load_workbook(file_bytes)
    ws = wb.active
    
    # 2. USE PANDAS FOR DATA LOGIC
    file_bytes.seek(0)
    df_preview = pd.read_excel(file_bytes, header=None, nrows=20)
    header_idx = 0
    for i, row in df_preview.iterrows():
        if any("sl no." in str(val).lower() for val in row.values):
            header_idx = i
            break
            
    file_bytes.seek(0)
    df_raw = pd.read_excel(file_bytes, header=header_idx)
    df_raw.columns = [str(c).strip() for c in df_raw.columns]
    
    u_cols = [c for c in df_raw.columns if c.endswith('_U')]

    # 3. APPLY LOGIC & WRITE TO CELLS
    for u_col in u_cols:
        base_name = u_col[:-2]
        l_col = base_name + '_L'
        
        u_col_idx = df_raw.columns.get_loc(u_col) + 1
        l_col_idx = df_raw.columns.get_loc(l_col) + 1 if l_col in df_raw.columns else None
        
        if l_col_idx:
            merged = df_raw[u_col].fillna(df_raw[l_col])
        else:
            merged = df_raw[u_col]
            
        numeric_series = pd.to_numeric(merged, errors='coerce')
        is_shutdown = merged.astype(str).str.contains("Site Shutdown", case=False, na=False)
        is_missing = numeric_series.isna() & ~is_shutdown
        
        global_mean = numeric_series.dropna().mean()
        if np.isnan(global_mean): global_mean = 20.0

        for idx in range(len(df_raw)):
            excel_row = idx + header_idx + 2
            
            if is_missing.iloc[idx]:
                window = numeric_series.iloc[max(0, idx-50):idx].dropna()
                if window.empty:
                    window = numeric_series.iloc[idx+1:idx+51].dropna()
                
                if not window.empty:
                    val = np.random.choice(window.values)
                    new_val = np.round(val * np.random.uniform(0.98, 1.02), 2)
                else:
                    new_val = np.round(global_mean * np.random.uniform(0.95, 1.05), 2)
                
                target_cell = ws.cell(row=excel_row, column=u_col_idx)
                target_cell.value = new_val
                target_cell.alignment = Alignment(horizontal='center', vertical='center')

        # 4. MERGE HEADERS
        if l_col_idx:
            ws.merge_cells(start_row=header_idx + 1, start_column=u_col_idx, 
                           end_row=header_idx + 1, end_column=l_col_idx)
            header_cell = ws.cell(row=header_idx + 1, column=u_col_idx)
            header_cell.value = base_name
            header_cell.alignment = Alignment(horizontal='center', vertical='center')

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# --- SAI GOUD BRANDED UI ---
st.title("🛡️ Sai Goud Report Processor")
st.info("Welcome to my page! Upload your pollution report in Excel format and I will handle the rest.")

# Sidebar Branding
st.sidebar.markdown("### User Controls")
file = st.sidebar.file_uploader("Upload Pollution Excel", type="xlsx")
st.sidebar.info("Developed by Sai Goud")

if file:
    if st.button("Press for Formatting"):
        with st.spinner("Processing report with Sai Goud's Logic..."):
            processed_file = process_pollution_report(file)
            
            st.success("Processing complete! Thank you for using Sai Goud's Website. Please download your processed report below and come back for more!")
            
            st.download_button(
                label="Download Processed Report",
                data=processed_file,
                file_name="Universal_Processed_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
